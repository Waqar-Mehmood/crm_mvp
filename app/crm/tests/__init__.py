import csv
import importlib
import json
import shutil
import tempfile
from datetime import timedelta
from io import BytesIO
from pathlib import Path
from unittest.mock import Mock, patch

from django.apps import apps as django_apps
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from PIL import Image
import requests

from crm.auth import (
    CRM_ROLE_ORDER,
    ROLE_MANAGER,
    ROLE_OWNER,
    ROLE_STAFF,
    ROLE_TEAM_LEAD,
    assign_crm_role,
    get_admin_assignable_role_choices,
    get_user_crm_role,
    get_user_role_status,
    user_has_minimum_crm_role,
    user_has_valid_crm_role,
)
from crm.models import (
    Company,
    CompanyEmail,
    CompanyPhone,
    CompanySocialLink,
    Contact,
    ContactEmail,
    ContactPhone,
    ContactSocialLink,
    ImportFile,
    ImportRow,
    SiteBranding,
)
from crm.services.google_sheets import (
    build_csv_export_url,
    extract_gid,
    extract_sheet_id,
    fetch_google_sheet_rows,
)
from crm.services.import_workflow import build_import_result_summary
from crm.services.import_parsers import (
    parse_csv_file,
    parse_google_sheet,
    parse_json_file,
    parse_xlsx_file,
)
from crm.services.import_service import (
    get_row_headers,
    parse_rows_from_source,
    rows_to_temporary_csv,
    rows_to_uploaded_csv,
    select_import_parser,
)


def make_logo_file(name="logo.png"):
    buffer = BytesIO()
    Image.new("RGBA", (1, 1), (209, 125, 47, 255)).save(buffer, format="PNG")
    return SimpleUploadedFile(name, buffer.getvalue(), content_type="image/png")


def make_csv_file(name="contacts.csv", content="Company Name,Email\nAcme Labs,person@example.com\n"):
    return SimpleUploadedFile(name, content.encode("utf-8"), content_type="text/csv")


def make_xlsx_file(name="contacts.xlsx", rows=None):
    buffer = BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows or (
        ["Company Name", "Email"],
        ["Acme Labs", "person@example.com"],
    ):
        worksheet.append(row)
    workbook.save(buffer)
    workbook.close()
    return SimpleUploadedFile(
        name,
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def make_json_file(name="contacts.json", rows=None):
    payload = rows or [{"Company Name": "Acme Labs", "Email": "person@example.com"}]
    return SimpleUploadedFile(
        name,
        json.dumps(payload).encode("utf-8"),
        content_type="application/json",
    )


class CRMRoleTestMixin:
    default_password = "test-pass-123"

    def create_user(self, username, role=None, password=None, **extra_fields):
        user = get_user_model().objects.create_user(
            username=username,
            password=password or self.default_password,
            **extra_fields,
        )
        if role:
            assign_crm_role(user, role)
            user.refresh_from_db()
        return user

    def create_superuser(self, username="superadmin", password=None, **extra_fields):
        return get_user_model().objects.create_superuser(
            username=username,
            password=password or self.default_password,
            **extra_fields,
        )


class AdvancedFilterTestMixin(CRMRoleTestMixin):
    def setUp(self):
        self.client = Client()
        self.staff_user = self.create_user("staffer", role=ROLE_STAFF)
        self.client.force_login(self.staff_user)

        self.acme = Company.objects.create(
            name="Acme Labs",
            industry="SaaS",
            company_size="12",
            revenue="$1M",
            address="1 Market Street",
            city="San Francisco",
            state="CA",
            country="US",
            notes="Outbound team",
        )
        self.brick = Company.objects.create(
            name="Brick Health",
            industry="Medical Facilities",
            company_size="4",
            city="Austin",
            state="TX",
            country="US",
        )
        self.cedar = Company.objects.create(
            name="Cedar Staffing",
            industry="Staffing and Recruiting",
            company_size="30",
            revenue="$5M",
            city="Los Angeles",
            state="CA",
            country="US",
        )

        CompanyPhone.objects.create(company=self.acme, phone="111-111", label="office")
        CompanyEmail.objects.create(company=self.acme, email="hello@acme.com", label="sales")
        CompanySocialLink.objects.create(company=self.acme, platform="linkedin", url="https://example.com/acme")
        CompanyEmail.objects.create(company=self.cedar, email="team@cedar.com", label="support")

        self.alice = Contact.objects.create(
            full_name="Alice Johnson",
            title="Growth Marketing Manager",
            notes="Works the SaaS pipeline",
        )
        self.bob = Contact.objects.create(
            full_name="Bob Stone",
            title="Director",
            notes="No linked company",
        )
        self.carla = Contact.objects.create(
            full_name="Carla Recruiter",
            title="Owner / Recruiter",
        )

        ContactEmail.objects.create(contact=self.alice, email="alice@acme.com", label="work", is_primary=True)
        ContactPhone.objects.create(contact=self.alice, phone="555-0101", label="work", is_primary=True)
        ContactEmail.objects.create(contact=self.carla, email="carla@example.com", label="work", is_primary=True)
        ContactPhone.objects.create(contact=self.carla, phone="555-0199", label="work", is_primary=True)
        ContactSocialLink.objects.create(contact=self.alice, platform="linkedin", url="https://example.com/alice")
        self.acme.contacts.add(self.alice)
        self.cedar.contacts.add(self.carla)

        base_time = timezone.now()
        self._set_created_at(self.acme, base_time - timedelta(days=12))
        self._set_created_at(self.brick, base_time - timedelta(days=5))
        self._set_created_at(self.cedar, base_time - timedelta(days=1))
        self._set_created_at(self.alice, base_time - timedelta(days=11))
        self._set_created_at(self.bob, base_time - timedelta(days=4))
        self._set_created_at(self.carla, base_time - timedelta(days=1))

        for index in range(12):
            company = Company.objects.create(
                name=f"California Extra {index + 1}",
                industry="SaaS",
                company_size="8",
                city="San Diego",
                state="CA",
                country="US",
            )
            self._set_created_at(company, base_time - timedelta(days=20))

        for index in range(12):
            contact = Contact.objects.create(
                full_name=f"Analyst {index + 1}",
                title="Analyst",
            )
            self.acme.contacts.add(contact)
            self._set_created_at(contact, base_time - timedelta(days=15))

    def _set_created_at(self, obj, created_at):
        obj.__class__.objects.filter(pk=obj.pk).update(created_at=created_at)
        obj.refresh_from_db()

    def parse_csv_response(self, response):
        rows = list(csv.reader(response.content.decode("utf-8").splitlines()))
        return rows[0], rows[1:]

    def parse_xlsx_response(self, response):
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        return rows[0], rows[1:]
