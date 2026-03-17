import csv
from io import BytesIO, StringIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

COMPANY_EXPORT_COLUMNS = (
    ("id", "ID"),
    ("company_name", "Company Name"),
    ("industry", "Industry"),
    ("company_size", "Company Size"),
    ("revenue", "Revenue"),
    ("address", "Address"),
    ("city", "City"),
    ("state", "State"),
    ("zip_code", "Zip Code"),
    ("country", "Country"),
    ("notes", "Notes"),
    ("phones", "Phones"),
    ("emails", "Emails"),
    ("profiles", "Profiles"),
    ("created_at", "Created At"),
)

CONTACT_EXPORT_COLUMNS = (
    ("id", "ID"),
    ("full_name", "Full Name"),
    ("title", "Title"),
    ("notes", "Notes"),
    ("primary_email", "Primary Email"),
    ("primary_phone", "Primary Phone"),
    ("emails", "Emails"),
    ("phones", "Phones"),
    ("companies", "Companies"),
    ("profiles", "Profiles"),
    ("created_at", "Created At"),
)

EXPORT_DATETIME_FORMAT = "%Y-%m-%d %H:%M"
XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def build_export_filename(base_name, extension):
    timestamp = timezone.localtime().strftime("%Y%m%d-%H%M%S")
    return f"{base_name}-export-{timestamp}.{extension}"


def export_rows_to_csv_response(base_name, columns, rows):
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow([label for _key, label in columns])
    for row in rows:
        writer.writerow([row.get(key, "") for key, _label in columns])

    response = HttpResponse(
        buffer.getvalue(),
        content_type="text/csv; charset=utf-8",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{build_export_filename(base_name, "csv")}"'
    )
    return response


def export_rows_to_xlsx_response(base_name, sheet_name, columns, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.freeze_panes = "A2"

    headers = [label for _key, label in columns]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([row.get(key, "") for key, _label in columns])

    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            max_length = max(max_length, len(str(cell.value or "")))
        worksheet.column_dimensions[column_letter].width = max(12, min(max_length + 2, 48))

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = (
        f'attachment; filename="{build_export_filename(base_name, "xlsx")}"'
    )
    return response


def format_export_datetime(value):
    if not value:
        return ""
    return timezone.localtime(value).strftime(EXPORT_DATETIME_FORMAT)


def join_export_values(values):
    return "; ".join(value for value in values if value)


def format_labeled_value(label, value):
    label = (label or "").strip()
    value = (value or "").strip()
    if not value:
        return ""
    if label:
        return f"{label.title()}: {value}"
    return value


def format_profile_value(profile):
    platform = (profile.platform or "").strip()
    url = (profile.url or "").strip()
    if not url:
        return ""
    if platform:
        return f"{platform.title()}: {url}"
    return url


def serialize_company_export_row(company):
    return {
        "id": company.id,
        "company_name": company.name,
        "industry": company.industry,
        "company_size": company.company_size,
        "revenue": company.revenue,
        "address": company.address,
        "city": company.city,
        "state": company.state,
        "zip_code": company.zip_code,
        "country": company.country,
        "notes": company.notes,
        "phones": join_export_values(
            format_labeled_value(phone.label, phone.phone)
            for phone in company.phones.all()
        ),
        "emails": join_export_values(
            format_labeled_value(email.label, email.email)
            for email in company.emails.all()
        ),
        "profiles": join_export_values(
            format_profile_value(link)
            for link in company.social_links.all()
        ),
        "created_at": format_export_datetime(company.created_at),
    }


def serialize_contact_export_row(contact):
    return {
        "id": contact.id,
        "full_name": contact.full_name,
        "title": contact.title,
        "notes": contact.notes,
        "primary_email": contact.email,
        "primary_phone": contact.phone,
        "emails": join_export_values(
            format_labeled_value(email.label, email.email)
            for email in contact.emails.all()
        ),
        "phones": join_export_values(
            format_labeled_value(phone.label, phone.phone)
            for phone in contact.phones.all()
        ),
        "companies": join_export_values(
            company.name
            for company in contact.companies.all()
        ),
        "profiles": join_export_values(
            format_profile_value(link)
            for link in contact.social_links.all()
        ),
        "created_at": format_export_datetime(contact.created_at),
    }
