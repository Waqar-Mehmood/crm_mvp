import csv
import importlib
from io import BytesIO
from pathlib import Path
import shutil
import tempfile
from datetime import timedelta
from unittest.mock import Mock, patch

from django.apps import apps as django_apps
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image
import requests

from crm.auth import (
    CRM_ROLE_ORDER,
    ROLE_MANAGER,
    ROLE_OWNER,
    ROLE_STAFF,
    ROLE_TEAM_LEAD,
    assign_crm_role,
    get_admin_assignable_role_choices,
    get_user_crm_role,
    get_user_role_status,
    user_has_minimum_crm_role,
    user_has_valid_crm_role,
)
from crm.models import (
    Company,
    CompanyEmail,
    CompanyPhone,
    CompanySocialLink,
    Contact,
    ContactEmail,
    ContactPhone,
    ContactSocialLink,
    ImportFile,
    ImportRow,
    SiteBranding,
)
from crm.services.google_sheets import (
    build_csv_export_url,
    extract_gid,
    extract_sheet_id,
    fetch_google_sheet_rows,
)
from crm.services.import_service import (
    get_row_headers,
    rows_to_temporary_csv,
    rows_to_uploaded_csv,
)


def make_logo_file(name="logo.png"):
    buffer = BytesIO()
    Image.new("RGBA", (1, 1), (209, 125, 47, 255)).save(buffer, format="PNG")
    return SimpleUploadedFile(name, buffer.getvalue(), content_type="image/png")


def make_csv_file(name="contacts.csv", content="Company Name,Email\nAcme Labs,person@example.com\n"):
    return SimpleUploadedFile(name, content.encode("utf-8"), content_type="text/csv")


class CRMRoleTestMixin:
    default_password = "test-pass-123"

    def create_user(self, username, role=None, password=None, **extra_fields):
        user = get_user_model().objects.create_user(
            username=username,
            password=password or self.default_password,
            **extra_fields,
        )
        if role:
            assign_crm_role(user, role)
            user.refresh_from_db()
        return user

    def create_superuser(self, username="superadmin", password=None, **extra_fields):
        return get_user_model().objects.create_superuser(
            username=username,
            password=password or self.default_password,
            **extra_fields,
        )


class CRMRoleHelperTests(CRMRoleTestMixin, TestCase):
    def test_crm_groups_exist_after_migration(self):
        self.assertEqual(Group.objects.filter(name__in=CRM_ROLE_ORDER).count(), len(CRM_ROLE_ORDER))

    def test_single_role_resolves_correctly(self):
        user = self.create_user("staff-user", role=ROLE_STAFF)

        self.assertEqual(get_user_crm_role(user), ROLE_STAFF)
        self.assertEqual(get_user_role_status(user), "valid")
        self.assertTrue(user_has_valid_crm_role(user))
        self.assertFalse(user.is_staff)

    def test_user_with_no_role_is_invalid(self):
        user = self.create_user("no-role")

        self.assertIsNone(get_user_crm_role(user))
        self.assertEqual(get_user_role_status(user), "missing")
        self.assertFalse(user_has_valid_crm_role(user))
        self.assertFalse(user_has_minimum_crm_role(user, ROLE_STAFF))

    def test_user_with_multiple_roles_is_invalid(self):
        user = self.create_user("multi-role")
        user.groups.add(
            Group.objects.get(name=ROLE_STAFF),
            Group.objects.get(name=ROLE_TEAM_LEAD),
        )
        user.refresh_from_db()

        self.assertIsNone(get_user_crm_role(user))
        self.assertEqual(get_user_role_status(user), "multiple")
        self.assertFalse(user_has_valid_crm_role(user))
        self.assertFalse(user.is_staff)

    def test_superuser_bypasses_crm_role_validation(self):
        user = self.create_superuser()

        self.assertTrue(user_has_valid_crm_role(user))
        self.assertEqual(get_user_role_status(user), "superuser")
        self.assertTrue(user_has_minimum_crm_role(user, ROLE_OWNER))
        self.assertTrue(user.is_staff)

    def test_assigning_and_switching_roles_syncs_staff_flag(self):
        user = self.create_user("role-switcher")

        assign_crm_role(user, ROLE_MANAGER)
        user.refresh_from_db()
        self.assertEqual(get_user_crm_role(user), ROLE_MANAGER)
        self.assertTrue(user.is_staff)

        assign_crm_role(user, ROLE_STAFF)
        user.refresh_from_db()
        self.assertEqual(get_user_crm_role(user), ROLE_STAFF)
        self.assertFalse(user.is_staff)

    def test_manager_assignable_role_choices_are_limited(self):
        user = self.create_user("manager-user", role=ROLE_MANAGER)

        assignable_roles = [role_name for role_name, _label in get_admin_assignable_role_choices(user)]

        self.assertEqual(assignable_roles, [ROLE_STAFF, ROLE_TEAM_LEAD])


class FrontendRoleAccessTests(CRMRoleTestMixin, TestCase):
    def setUp(self):
        self.client = Client()
        self.staff_user = self.create_user("staffer", role=ROLE_STAFF)
        self.team_lead_user = self.create_user("teamlead", role=ROLE_TEAM_LEAD)
        self.manager_user = self.create_user("manager", role=ROLE_MANAGER)
        self.owner_user = self.create_user("owner", role=ROLE_OWNER)
        self.invalid_user = self.create_user("invalid-user")
        self.superuser = self.create_superuser(username="waqar")

        Company.objects.create(name="Acme Labs")
        Contact.objects.create(full_name="Jane Example")
        self.import_file = ImportFile.objects.create(file_name="seed.csv")
        ImportRow.objects.create(
            import_file=self.import_file,
            row_number=1,
            company_name="Acme Labs",
            contact_name="Jane Example",
        )

    def prime_import_session(self):
        session = self.client.session
        session["import_csv_temp_path"] = "/tmp/test-import.csv"
        session["import_csv_original_name"] = "test-import.csv"
        session["import_csv_headers"] = ["Company Name", "Email"]
        session.save()

    def test_anonymous_users_are_redirected_to_login(self):
        protected_urls = [
            reverse("home"),
            reverse("company_list"),
            reverse("contact_list"),
            reverse("import_file_list"),
            reverse("import_file_detail", args=[self.import_file.pk]),
            reverse("import_upload"),
            reverse("import_map_headers"),
        ]

        for url in protected_urls:
            with self.subTest(url=url):
                response = self.client.get(url)
                self.assertEqual(response.status_code, 302)
                self.assertIn(f"{reverse('login')}?next=", response["Location"])

    def test_anonymous_export_requests_are_redirected_to_login(self):
        for url in (reverse("company_list"), reverse("contact_list")):
            with self.subTest(url=url):
                response = self.client.get(url, {"export": "csv"})
                self.assertEqual(response.status_code, 302)
                self.assertIn(f"{reverse('login')}?next=", response["Location"])

    def test_staff_can_browse_but_not_access_import_mutations(self):
        self.client.login(username="staffer", password=self.default_password)

        for url in [
            reverse("company_list"),
            reverse("contact_list"),
            reverse("import_file_list"),
            reverse("import_file_detail", args=[self.import_file.pk]),
        ]:
            with self.subTest(url=url):
                response = self.client.get(url)
                self.assertEqual(response.status_code, 200)

        for url in [reverse("import_upload"), reverse("import_map_headers")]:
            with self.subTest(url=url):
                response = self.client.get(url)
                self.assertEqual(response.status_code, 403)

    def test_team_lead_and_above_can_access_import_pages(self):
        for username in ("teamlead", "manager", "owner", "waqar"):
            with self.subTest(username=username):
                self.client.force_login(get_user_model().objects.get(username=username))
                upload_response = self.client.get(reverse("import_upload"))
                self.assertEqual(upload_response.status_code, 200)

                self.prime_import_session()
                mapping_response = self.client.get(reverse("import_map_headers"))
                self.assertEqual(mapping_response.status_code, 200)
                self.client.logout()

    def test_invalid_role_user_gets_forbidden_on_crm_pages(self):
        self.client.login(username="invalid-user", password=self.default_password)

        browse_response = self.client.get(reverse("company_list"))
        upload_response = self.client.get(reverse("import_upload"))

        self.assertEqual(browse_response.status_code, 403)
        self.assertEqual(upload_response.status_code, 403)

    def test_multiple_role_user_gets_forbidden_on_crm_pages(self):
        user = self.create_user("broken-user", role=ROLE_STAFF)
        user.groups.add(Group.objects.get(name=ROLE_MANAGER))
        self.client.login(username="broken-user", password=self.default_password)

        response = self.client.get(reverse("contact_list"))

        self.assertEqual(response.status_code, 403)

    def test_upload_navigation_is_role_aware(self):
        self.client.login(username="staffer", password=self.default_password)
        staff_response = self.client.get(reverse("import_file_list"))
        self.assertEqual(staff_response.status_code, 200)
        self.assertNotContains(staff_response, 'href="/imports/upload/"')
        self.client.logout()

        self.client.login(username="teamlead", password=self.default_password)
        lead_response = self.client.get(reverse("import_file_list"))
        self.assertEqual(lead_response.status_code, 200)
        self.assertContains(lead_response, 'href="/imports/upload/"')

    def test_login_redirects_authenticated_users_to_companies(self):
        self.client.login(username="staffer", password=self.default_password)

        response = self.client.get(reverse("login"))

        self.assertRedirects(response, reverse("company_list"))

    def test_login_uses_next_parameter(self):
        response = self.client.post(
            reverse("login"),
            {
                "username": "staffer",
                "password": self.default_password,
                "next": reverse("contact_list"),
            },
        )

        self.assertRedirects(response, reverse("contact_list"))

    def test_logout_redirects_to_login(self):
        self.client.login(username="staffer", password=self.default_password)

        response = self.client.post(reverse("logout"))

        self.assertRedirects(response, reverse("login"))

    def test_login_template_renders_branding(self):
        response = self.client.get(reverse("login"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "The Zulfis CRM")


class AdminAccessTests(CRMRoleTestMixin, TestCase):
    def setUp(self):
        self.client = Client()
        self.staff_user = self.create_user("staffer", role=ROLE_STAFF)
        self.team_lead_user = self.create_user("teamlead", role=ROLE_TEAM_LEAD)
        self.manager_user = self.create_user("manager", role=ROLE_MANAGER)
        self.owner_user = self.create_user("owner", role=ROLE_OWNER)
        self.superuser = self.create_superuser(username="superadmin")

    def test_staff_and_team_lead_cannot_access_admin(self):
        for username in ("staffer", "teamlead"):
            with self.subTest(username=username):
                self.client.login(username=username, password=self.default_password)
                response = self.client.get(reverse("admin:auth_user_changelist"))
                self.assertEqual(response.status_code, 302)
                self.assertIn(reverse("admin:login"), response["Location"])
                self.client.logout()

    def test_manager_owner_and_superuser_can_access_admin(self):
        for username in ("manager", "owner", "superadmin"):
            with self.subTest(username=username):
                self.client.login(username=username, password=self.default_password)
                index_response = self.client.get(reverse("admin:index"))
                changelist_response = self.client.get(reverse("admin:auth_user_changelist"))
                self.assertEqual(index_response.status_code, 200)
                self.assertEqual(changelist_response.status_code, 200)
                self.client.logout()


class UserAdminRoleManagementTests(CRMRoleTestMixin, TestCase):
    def setUp(self):
        self.client = Client()
        self.admin_user = self.create_superuser(username="superadmin")
        self.target_user = self.create_user("target-user")
        self.client.force_login(self.admin_user)

    def update_user_role_via_admin(self, role_name):
        return self.client.post(
            reverse("admin:auth_user_change", args=[self.target_user.pk]),
            {
                "username": self.target_user.username,
                "first_name": "",
                "last_name": "",
                "email": "",
                "crm_role": role_name,
                "is_active": "on",
                "_save": "Save",
            },
            follow=True,
        )

    def test_assigning_manager_role_via_admin_updates_staff_status(self):
        response = self.update_user_role_via_admin(ROLE_MANAGER)
        self.target_user.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(get_user_crm_role(self.target_user), ROLE_MANAGER)
        self.assertTrue(self.target_user.is_staff)

    def test_switching_manager_to_staff_via_admin_removes_admin_access(self):
        self.update_user_role_via_admin(ROLE_MANAGER)
        self.update_user_role_via_admin(ROLE_STAFF)
        self.target_user.refresh_from_db()

        self.assertEqual(get_user_crm_role(self.target_user), ROLE_STAFF)
        self.assertFalse(self.target_user.is_staff)


class ManagerScopedAdminRestrictionsTests(CRMRoleTestMixin, TestCase):
    def setUp(self):
        self.client = Client()
        self.manager_user = self.create_user("manager", role=ROLE_MANAGER)
        self.owner_user = self.create_user("owner", role=ROLE_OWNER)
        self.peer_manager = self.create_user("peer-manager", role=ROLE_MANAGER)
        self.staff_user = self.create_user("staffer", role=ROLE_STAFF)
        self.team_lead_user = self.create_user("teamlead", role=ROLE_TEAM_LEAD)
        self.superuser = self.create_superuser(username="superadmin")
        self.site_branding = SiteBranding.objects.create(site_name="The Zulfis CRM")

    def force_login(self, user):
        self.client.force_login(user)

    def admin_change_url(self, user):
        return reverse("admin:auth_user_change", args=[user.pk])

    def admin_delete_url(self, user):
        return reverse("admin:auth_user_delete", args=[user.pk])

    def admin_password_url(self, user):
        return reverse("admin:auth_user_password_change", args=[user.pk])

    def extract_role_choices(self, response):
        return [
            choice_value
            for choice_value, _choice_label in response.context["adminform"].form.fields["crm_role"].choices
            if choice_value
        ]

    def editable_field_names(self, response):
        return set(response.context["adminform"].form.fields.keys())

    def app_model_names(self, response, app_label):
        for app in response.context["app_list"]:
            if app["app_label"] == app_label:
                return {model["object_name"] for model in app["models"]}
        return set()

    def test_manager_add_form_shows_only_lower_roles(self):
        self.force_login(self.manager_user)

        response = self.client.get(reverse("admin:auth_user_add"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.extract_role_choices(response), [ROLE_STAFF, ROLE_TEAM_LEAD])

    def test_owner_and_superuser_add_forms_show_all_roles(self):
        for actor in (self.owner_user, self.superuser):
            with self.subTest(actor=actor.username):
                self.force_login(actor)
                response = self.client.get(reverse("admin:auth_user_add"))
                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    self.extract_role_choices(response),
                    [ROLE_STAFF, ROLE_TEAM_LEAD, ROLE_MANAGER, ROLE_OWNER],
                )
                self.client.logout()

    def test_manager_change_form_for_lower_role_shows_only_lower_roles(self):
        self.force_login(self.manager_user)

        response = self.client.get(self.admin_change_url(self.staff_user))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["has_change_permission"], True)
        self.assertEqual(self.extract_role_choices(response), [ROLE_STAFF, ROLE_TEAM_LEAD])

    def test_manager_can_reassign_lower_role_user_to_team_lead(self):
        self.force_login(self.manager_user)

        response = self.client.post(
            self.admin_change_url(self.staff_user),
            {
                "username": self.staff_user.username,
                "first_name": "",
                "last_name": "",
                "email": "",
                "crm_role": ROLE_TEAM_LEAD,
                "is_active": "on",
                "_save": "Save",
            },
            follow=True,
        )
        self.staff_user.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(get_user_crm_role(self.staff_user), ROLE_TEAM_LEAD)

    def test_manager_cannot_escalate_lower_role_user_to_manager(self):
        self.force_login(self.manager_user)

        response = self.client.post(
            self.admin_change_url(self.staff_user),
            {
                "username": self.staff_user.username,
                "first_name": "",
                "last_name": "",
                "email": "",
                "crm_role": ROLE_MANAGER,
                "is_active": "on",
                "_save": "Save",
            },
        )
        self.staff_user.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Select a valid choice")
        self.assertEqual(get_user_crm_role(self.staff_user), ROLE_STAFF)

    def test_manager_changelist_shows_owner_and_managers_but_hides_superusers(self):
        self.force_login(self.manager_user)

        response = self.client.get(reverse("admin:auth_user_changelist"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.manager_user.username)
        self.assertContains(response, self.peer_manager.username)
        self.assertContains(response, self.owner_user.username)
        self.assertContains(response, self.staff_user.username)
        self.assertNotContains(response, self.superuser.username)

    def test_manager_can_open_owner_and_manager_records_in_read_only_mode(self):
        self.force_login(self.manager_user)

        for target in (self.peer_manager, self.owner_user):
            with self.subTest(target=target.username):
                response = self.client.get(self.admin_change_url(target))
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.context["has_change_permission"], False)
                self.assertNotContains(response, 'name="_save"')

    def test_manager_can_open_self_with_profile_only_edit_access(self):
        self.force_login(self.manager_user)

        response = self.client.get(self.admin_change_url(self.manager_user))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["has_change_permission"], True)
        self.assertContains(response, 'name="_save"')
        self.assertEqual(
            self.editable_field_names(response),
            {"first_name", "last_name", "email", "password"},
        )

    def test_owner_can_open_self_with_profile_only_edit_access(self):
        self.force_login(self.owner_user)

        response = self.client.get(self.admin_change_url(self.owner_user))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["has_change_permission"], True)
        self.assertContains(response, 'name="_save"')
        self.assertEqual(
            self.editable_field_names(response),
            {"first_name", "last_name", "email", "password"},
        )

    def test_manager_cannot_edit_disallowed_targets_via_post(self):
        self.force_login(self.manager_user)

        for target in (self.peer_manager, self.owner_user):
            with self.subTest(target=target.username):
                response = self.client.post(
                    self.admin_change_url(target),
                    {
                        "username": target.username,
                        "first_name": "",
                        "last_name": "",
                        "email": "",
                        "crm_role": ROLE_STAFF,
                        "is_active": "on",
                        "_save": "Save",
                    },
                )
                self.assertEqual(response.status_code, 403)

    def test_manager_can_update_own_profile_without_changing_access_fields(self):
        self.force_login(self.manager_user)

        response = self.client.post(
            self.admin_change_url(self.manager_user),
            {
                "username": "hacked-manager",
                "first_name": "Rizwan",
                "last_name": "Manager",
                "email": "manager@example.com",
                "crm_role": ROLE_OWNER,
                "_save": "Save",
            },
            follow=True,
        )
        self.manager_user.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.manager_user.username, "manager")
        self.assertEqual(self.manager_user.first_name, "Rizwan")
        self.assertEqual(self.manager_user.last_name, "Manager")
        self.assertEqual(self.manager_user.email, "manager@example.com")
        self.assertEqual(get_user_crm_role(self.manager_user), ROLE_MANAGER)
        self.assertTrue(self.manager_user.is_active)

    def test_owner_can_update_own_profile_without_changing_access_fields(self):
        self.force_login(self.owner_user)

        response = self.client.post(
            self.admin_change_url(self.owner_user),
            {
                "username": "hacked-owner",
                "first_name": "Owner",
                "last_name": "Profile",
                "email": "owner@example.com",
                "crm_role": ROLE_STAFF,
                "_save": "Save",
            },
            follow=True,
        )
        self.owner_user.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.owner_user.username, "owner")
        self.assertEqual(self.owner_user.first_name, "Owner")
        self.assertEqual(self.owner_user.last_name, "Profile")
        self.assertEqual(self.owner_user.email, "owner@example.com")
        self.assertEqual(get_user_crm_role(self.owner_user), ROLE_OWNER)
        self.assertTrue(self.owner_user.is_active)

    def test_manager_cannot_delete_self_or_higher_role_users(self):
        self.force_login(self.manager_user)

        for target in (self.manager_user, self.peer_manager, self.owner_user):
            with self.subTest(target=target.username):
                response = self.client.post(self.admin_delete_url(target), {"post": "yes"})
                self.assertEqual(response.status_code, 403)

    def test_owner_cannot_delete_self(self):
        self.force_login(self.owner_user)

        response = self.client.post(self.admin_delete_url(self.owner_user), {"post": "yes"})

        self.assertEqual(response.status_code, 403)

    def test_manager_can_access_password_change_for_self_and_lower_roles(self):
        self.force_login(self.manager_user)

        for target in (self.manager_user, self.staff_user, self.team_lead_user):
            with self.subTest(target=target.username):
                response = self.client.get(self.admin_password_url(target))
                self.assertEqual(response.status_code, 200)

    def test_manager_cannot_access_password_change_for_manager_or_above(self):
        self.force_login(self.manager_user)

        for target in (self.peer_manager, self.owner_user, self.superuser):
            with self.subTest(target=target.username):
                response = self.client.get(self.admin_password_url(target))
                self.assertEqual(response.status_code, 403)

    def test_manager_can_change_own_password(self):
        self.force_login(self.manager_user)

        response = self.client.post(
            self.admin_password_url(self.manager_user),
            {
                "password1": "manager-new-pass-123",
                "password2": "manager-new-pass-123",
            },
            follow=True,
        )
        self.manager_user.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(self.manager_user.check_password("manager-new-pass-123"))

    def test_owner_can_change_own_password(self):
        self.force_login(self.owner_user)

        response = self.client.post(
            self.admin_password_url(self.owner_user),
            {
                "password1": "owner-new-pass-123",
                "password2": "owner-new-pass-123",
            },
            follow=True,
        )
        self.owner_user.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(self.owner_user.check_password("owner-new-pass-123"))

    def test_manager_can_delete_lower_role_user(self):
        self.force_login(self.manager_user)

        confirm_response = self.client.get(self.admin_delete_url(self.team_lead_user))
        delete_response = self.client.post(self.admin_delete_url(self.team_lead_user), {"post": "yes"}, follow=True)

        self.assertEqual(confirm_response.status_code, 200)
        self.assertEqual(delete_response.status_code, 200)
        self.assertFalse(get_user_model().objects.filter(pk=self.team_lead_user.pk).exists())

    def test_manager_cannot_view_site_branding_in_index_or_direct_url(self):
        self.force_login(self.manager_user)

        index_response = self.client.get(reverse("admin:index"))
        changelist_response = self.client.get(reverse("admin:crm_sitebranding_changelist"))

        self.assertEqual(index_response.status_code, 200)
        self.assertNotIn("SiteBranding", self.app_model_names(index_response, "crm"))
        self.assertEqual(changelist_response.status_code, 403)

    def test_owner_and_superuser_can_access_site_branding(self):
        for actor in (self.owner_user, self.superuser):
            with self.subTest(actor=actor.username):
                self.force_login(actor)
                index_response = self.client.get(reverse("admin:index"))
                changelist_response = self.client.get(reverse("admin:crm_sitebranding_changelist"))
                self.assertEqual(index_response.status_code, 200)
                self.assertIn("SiteBranding", self.app_model_names(index_response, "crm"))
                self.assertEqual(changelist_response.status_code, 200)
                self.client.logout()


class AdvancedFilterTests(CRMRoleTestMixin, TestCase):
    def setUp(self):
        self.client = Client()
        self.staff_user = self.create_user("staffer", role=ROLE_STAFF)
        self.client.force_login(self.staff_user)

        self.acme = Company.objects.create(
            name="Acme Labs",
            industry="SaaS",
            company_size="12",
            revenue="$1M",
            address="1 Market Street",
            city="San Francisco",
            state="CA",
            country="US",
            notes="Outbound team",
        )
        self.brick = Company.objects.create(
            name="Brick Health",
            industry="Medical Facilities",
            company_size="4",
            city="Austin",
            state="TX",
            country="US",
        )
        self.cedar = Company.objects.create(
            name="Cedar Staffing",
            industry="Staffing and Recruiting",
            company_size="30",
            revenue="$5M",
            city="Los Angeles",
            state="CA",
            country="US",
        )

        CompanyPhone.objects.create(company=self.acme, phone="111-111", label="office")
        CompanyEmail.objects.create(company=self.acme, email="hello@acme.com", label="sales")
        CompanySocialLink.objects.create(company=self.acme, platform="linkedin", url="https://example.com/acme")
        CompanyEmail.objects.create(company=self.cedar, email="team@cedar.com", label="support")

        self.alice = Contact.objects.create(
            full_name="Alice Johnson",
            title="Growth Marketing Manager",
            notes="Works the SaaS pipeline",
        )
        self.bob = Contact.objects.create(
            full_name="Bob Stone",
            title="Director",
            notes="No linked company",
        )
        self.carla = Contact.objects.create(
            full_name="Carla Recruiter",
            title="Owner / Recruiter",
            email="carla@example.com",
            phone="555-0199",
        )

        ContactEmail.objects.create(contact=self.alice, email="alice@acme.com", label="work")
        ContactPhone.objects.create(contact=self.alice, phone="555-0101", label="work")
        ContactSocialLink.objects.create(contact=self.alice, platform="linkedin", url="https://example.com/alice")
        self.acme.contacts.add(self.alice)
        self.cedar.contacts.add(self.carla)

        base_time = timezone.now()
        self._set_created_at(self.acme, base_time - timedelta(days=12))
        self._set_created_at(self.brick, base_time - timedelta(days=5))
        self._set_created_at(self.cedar, base_time - timedelta(days=1))
        self._set_created_at(self.alice, base_time - timedelta(days=11))
        self._set_created_at(self.bob, base_time - timedelta(days=4))
        self._set_created_at(self.carla, base_time - timedelta(days=1))

        for index in range(12):
            company = Company.objects.create(
                name=f"California Extra {index + 1}",
                industry="SaaS",
                company_size="8",
                city="San Diego",
                state="CA",
                country="US",
            )
            self._set_created_at(company, base_time - timedelta(days=20))

        for index in range(12):
            contact = Contact.objects.create(
                full_name=f"Analyst {index + 1}",
                title="Analyst",
            )
            self.acme.contacts.add(contact)
            self._set_created_at(contact, base_time - timedelta(days=15))

    def _set_created_at(self, obj, created_at):
        obj.__class__.objects.filter(pk=obj.pk).update(created_at=created_at)
        obj.refresh_from_db()

    def parse_csv_response(self, response):
        rows = list(csv.reader(response.content.decode("utf-8").splitlines()))
        return rows[0], rows[1:]

    def parse_xlsx_response(self, response):
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        return rows[0], rows[1:]

    def test_company_filters_narrow_results_and_keep_form_values(self):
        response = self.client.get(
            reverse("company_list"),
            {
                "industry": "SaaS",
                "state": "CA",
                "size_min": "10",
                "has_email": "yes",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual({company.name for company in response.context["companies"]}, {"Acme Labs"})
        self.assertEqual(response.context["filters"]["industry"], "SaaS")
        self.assertEqual(response.context["filters"]["state"], "CA")
        self.assertEqual(response.context["filters"]["size_min"], "10")
        self.assertEqual(response.context["filters"]["has_email"], "yes")
        self.assertContains(response, "Matching companies")
        self.assertContains(response, "Industry:")

    def test_company_revenue_and_created_date_filters_work(self):
        response = self.client.get(
            reverse("company_list"),
            {
                "revenue": "$5M",
                "created_from": str((timezone.now() - timedelta(days=2)).date()),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual({company.name for company in response.context["companies"]}, {"Cedar Staffing"})

    def test_company_pagination_preserves_filters(self):
        response = self.client.get(reverse("company_list"), {"state": "CA", "page": 2})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["page_obj"].number, 2)
        self.assertEqual(response.context["page_query"], "state=CA")
        self.assertContains(response, "?page=1&state=CA")

    def test_company_filtered_empty_state_appears_when_no_results_match(self):
        response = self.client.get(reverse("company_list"), {"industry": "Nonexistent"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No companies matched the current filters.")
        self.assertNotContains(response, "No companies have landed in the ledger.")

    def test_company_csv_export_uses_filtered_rows_and_full_columns(self):
        response = self.client.get(
            reverse("company_list"),
            {
                "industry": "SaaS",
                "has_email": "yes",
                "export": "csv",
            },
        )
        headers, rows = self.parse_csv_response(response)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn('attachment; filename="companies-export-', response["Content-Disposition"])
        self.assertEqual(
            headers,
            [
                "ID",
                "Company Name",
                "Industry",
                "Company Size",
                "Revenue",
                "Address",
                "City",
                "State",
                "Zip Code",
                "Country",
                "Notes",
                "Phones",
                "Emails",
                "Profiles",
                "Created At",
            ],
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][1], "Acme Labs")
        self.assertEqual(rows[0][10], "Outbound team")
        self.assertIn("Office: 111-111", rows[0][11])
        self.assertIn("Sales: hello@acme.com", rows[0][12])
        self.assertIn("Linkedin: https://example.com/acme", rows[0][13])

    def test_company_export_ignores_pagination_and_invalid_export_falls_back_to_html(self):
        export_response = self.client.get(
            reverse("company_list"),
            {"state": "CA", "page": 2, "export": "csv"},
        )
        _headers, rows = self.parse_csv_response(export_response)

        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(len(rows), 14)
        self.assertEqual({row[1] for row in rows}, {"Acme Labs", "Cedar Staffing", *{f"California Extra {index + 1}" for index in range(12)}})

        html_response = self.client.get(reverse("company_list"), {"export": "pdf"})
        self.assertEqual(html_response.status_code, 200)
        self.assertContains(html_response, "Company records")
        self.assertNotIn("Content-Disposition", html_response)

    def test_company_xlsx_export_and_empty_export_behave_correctly(self):
        response = self.client.get(
            reverse("company_list"),
            {
                "revenue": "$5M",
                "created_from": str((timezone.now() - timedelta(days=2)).date()),
                "export": "xlsx",
            },
        )
        headers, rows = self.parse_xlsx_response(response)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn('attachment; filename="companies-export-', response["Content-Disposition"])
        self.assertEqual(headers[1], "Company Name")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][1], "Cedar Staffing")
        self.assertEqual(rows[0][4], "$5M")

        empty_response = self.client.get(
            reverse("company_list"),
            {"industry": "Nonexistent", "export": "csv"},
        )
        empty_headers, empty_rows = self.parse_csv_response(empty_response)
        self.assertEqual(empty_headers[1], "Company Name")
        self.assertEqual(empty_rows, [])

    def test_contact_filters_match_related_data_and_presence_rules(self):
        response = self.client.get(
            reverse("contact_list"),
            {
                "q": "alice@acme.com",
                "company": "Acme",
                "has_profile": "yes",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual({contact.full_name for contact in response.context["contacts"]}, {"Alice Johnson"})
        self.assertEqual(response.context["filters"]["company"], "Acme")
        self.assertEqual(response.context["filters"]["has_profile"], "yes")

    def test_contact_created_and_presence_filters_work(self):
        response = self.client.get(
            reverse("contact_list"),
            {
                "has_company": "no",
                "has_email": "no",
                "has_phone": "no",
                "created_to": str((timezone.now() - timedelta(days=3)).date()),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual({contact.full_name for contact in response.context["contacts"]}, {"Bob Stone"})

    def test_contact_pagination_preserves_filters(self):
        response = self.client.get(reverse("contact_list"), {"title": "Analyst", "page": 2})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["page_obj"].number, 2)
        self.assertEqual(response.context["page_query"], "title=Analyst")
        self.assertContains(response, "?page=1&title=Analyst")

    def test_contact_filtered_empty_state_appears_when_no_results_match(self):
        response = self.client.get(reverse("contact_list"), {"title": "Astronaut"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "No contacts matched the current filters.")
        self.assertNotContains(response, "No contacts are available yet.")

    def test_contact_csv_export_uses_filtered_rows_and_joined_columns(self):
        response = self.client.get(
            reverse("contact_list"),
            {
                "q": "alice@acme.com",
                "company": "Acme",
                "has_profile": "yes",
                "export": "csv",
            },
        )
        headers, rows = self.parse_csv_response(response)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv; charset=utf-8")
        self.assertIn('attachment; filename="contacts-export-', response["Content-Disposition"])
        self.assertEqual(
            headers,
            [
                "ID",
                "Full Name",
                "Title",
                "Notes",
                "Primary Email",
                "Primary Phone",
                "Emails",
                "Phones",
                "Companies",
                "Profiles",
                "Created At",
            ],
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][1], "Alice Johnson")
        self.assertIn("Work: alice@acme.com", rows[0][6])
        self.assertIn("Work: 555-0101", rows[0][7])
        self.assertEqual(rows[0][8], "Acme Labs")
        self.assertIn("Linkedin: https://example.com/alice", rows[0][9])

    def test_contact_export_ignores_pagination_and_xlsx_returns_filtered_rows(self):
        csv_response = self.client.get(
            reverse("contact_list"),
            {"title": "Analyst", "page": 2, "export": "csv"},
        )
        _headers, csv_rows = self.parse_csv_response(csv_response)
        self.assertEqual(csv_response.status_code, 200)
        self.assertEqual(len(csv_rows), 12)

        xlsx_response = self.client.get(
            reverse("contact_list"),
            {
                "has_company": "no",
                "has_email": "no",
                "created_to": str((timezone.now() - timedelta(days=3)).date()),
                "export": "xlsx",
            },
        )
        headers, xlsx_rows = self.parse_xlsx_response(xlsx_response)
        self.assertEqual(headers[1], "Full Name")
        self.assertEqual(len(xlsx_rows), 1)
        self.assertEqual(xlsx_rows[0][1], "Bob Stone")


class GoogleSheetsServiceTests(TestCase):
    def test_extract_helpers_build_expected_export_url(self):
        sheet_url = (
            "https://docs.google.com/spreadsheets/d/"
            "1ngu9sB-ZtIFoA3BqBnd2AwBroIZL9_c_8ZdREHTe8NM/edit?gid=0#gid=0"
        )

        self.assertEqual(
            extract_sheet_id(sheet_url),
            "1ngu9sB-ZtIFoA3BqBnd2AwBroIZL9_c_8ZdREHTe8NM",
        )
        self.assertEqual(extract_gid(sheet_url), "0")
        self.assertEqual(
            build_csv_export_url(sheet_url),
            "https://docs.google.com/spreadsheets/d/"
            "1ngu9sB-ZtIFoA3BqBnd2AwBroIZL9_c_8ZdREHTe8NM/export?format=csv&gid=0",
        )

    def test_extract_gid_defaults_to_zero_when_missing(self):
        sheet_url = "https://docs.google.com/spreadsheets/d/test-sheet-id/edit"

        self.assertEqual(extract_gid(sheet_url), "0")

    @patch("crm.services.google_sheets.requests.get")
    def test_fetch_google_sheet_rows_parses_csv_into_dicts(self, mock_get):
        mock_response = Mock()
        mock_response.text = "Name,Email\nAlice,alice@example.com\nBob,bob@example.com\n"
        mock_response.raise_for_status.return_value = None
        mock_get.return_value = mock_response

        rows = fetch_google_sheet_rows(
            "https://docs.google.com/spreadsheets/d/test-sheet-id/edit?gid=0#gid=0"
        )

        self.assertEqual(
            rows,
            [
                {"Name": "Alice", "Email": "alice@example.com"},
                {"Name": "Bob", "Email": "bob@example.com"},
            ],
        )
        mock_get.assert_called_once()

    @patch("crm.services.google_sheets.requests.get")
    def test_fetch_google_sheet_rows_raises_runtime_error_on_request_failure(self, mock_get):
        mock_get.side_effect = requests.RequestException("network down")

        with self.assertRaises(RuntimeError):
            fetch_google_sheet_rows(
                "https://docs.google.com/spreadsheets/d/test-sheet-id/edit?gid=0#gid=0"
            )

    def test_fetch_google_sheet_rows_raises_value_error_for_invalid_url(self):
        with self.assertRaises(ValueError):
            fetch_google_sheet_rows("https://example.com/not-a-sheet")


class ImportServiceTests(TestCase):
    def test_get_row_headers_preserves_first_seen_order(self):
        rows = [
            {"Company Name": "Acme", "Email": "hello@acme.com"},
            {"Company Name": "Beta", "Phone": "555-0101"},
        ]

        self.assertEqual(get_row_headers(rows), ["Company Name", "Email", "Phone"])

    def test_rows_to_temporary_csv_writes_csv_file(self):
        rows = [
            {"Company Name": "Acme", "Email": "hello@acme.com"},
            {"Company Name": "Beta", "Phone": "555-0101"},
        ]

        temp_path = Path(rows_to_temporary_csv(rows))
        self.addCleanup(temp_path.unlink, missing_ok=True)

        self.assertTrue(temp_path.exists())
        self.assertEqual(
            temp_path.read_text(encoding="utf-8"),
            "Company Name,Email,Phone\nAcme,hello@acme.com,\nBeta,,555-0101\n",
        )

    def test_rows_to_uploaded_csv_returns_simple_uploaded_file(self):
        rows = [{"Company Name": "Acme", "Email": "hello@acme.com"}]

        uploaded = rows_to_uploaded_csv(rows)

        self.assertEqual(uploaded.name, "google_sheet_import.csv")
        self.assertEqual(uploaded.content_type, "text/csv")
        self.assertEqual(
            uploaded.read().decode("utf-8"),
            "Company Name,Email\r\nAcme,hello@acme.com\r\n",
        )

    def test_rows_to_uploaded_csv_rejects_empty_rows(self):
        with self.assertRaises(ValueError):
            rows_to_uploaded_csv([])


class GoogleSheetsImportFlowTests(CRMRoleTestMixin, TestCase):
    def setUp(self):
        self.client = Client()
        self.team_lead_user = self.create_user("teamlead", role=ROLE_TEAM_LEAD)
        self.client.force_login(self.team_lead_user)
        self.temp_root = Path(tempfile.mkdtemp())
        self.temp_base_dir = self.temp_root / "base"
        self.temp_media_root = self.temp_root / "media"
        self.temp_base_dir.mkdir(parents=True, exist_ok=True)
        self.temp_media_root.mkdir(parents=True, exist_ok=True)
        self.settings_override = override_settings(
            BASE_DIR=self.temp_base_dir,
            MEDIA_ROOT=self.temp_media_root,
            MEDIA_URL="/media/",
        )
        self.settings_override.enable()
        self.addCleanup(self.settings_override.disable)
        self.addCleanup(shutil.rmtree, self.temp_root, ignore_errors=True)

    def test_google_sheets_preview_page_renders(self):
        response = self.client.get(reverse("import_google_sheets"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Google Sheets URL")
        self.assertContains(response, "Preview sheet")

    @patch("crm.services.google_sheets.fetch_google_sheet_rows")
    def test_google_sheets_preview_post_shows_headers_and_first_rows(self, mock_fetch):
        mock_fetch.return_value = [
            {"Company Name": "Acme", "Email": "hello@acme.com"},
            {"Company Name": "Beta", "Email": "team@beta.com"},
            {"Company Name": "Cedar", "Email": "ops@cedar.com"},
        ]

        response = self.client.post(
            reverse("import_google_sheets"),
            {
                "sheet_url": "https://docs.google.com/spreadsheets/d/test-sheet/edit?gid=0#gid=0",
                "action": "preview",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["headers"], ["Company Name", "Email"])
        self.assertEqual(len(response.context["preview_rows"]), 3)
        self.assertEqual(response.context["total_rows"], 3)
        self.assertContains(response, "Continue to mapping")

    @patch("crm.services.google_sheets.fetch_google_sheet_rows")
    def test_google_sheets_import_action_reuses_mapping_session_flow(self, mock_fetch):
        mock_fetch.return_value = [
            {"Company Name": "Acme", "Email": "hello@acme.com"},
            {"Company Name": "Beta", "Email": "team@beta.com"},
        ]

        response = self.client.post(
            reverse("import_google_sheets"),
            {
                "sheet_url": "https://docs.google.com/spreadsheets/d/test-sheet/edit?gid=0#gid=0",
                "action": "import",
            },
        )

        self.assertRedirects(response, reverse("import_map_headers"))
        session = self.client.session
        temp_path = Path(session["import_csv_temp_path"])
        self.addCleanup(temp_path.unlink, missing_ok=True)
        self.assertTrue(temp_path.exists())
        self.assertEqual(session["import_csv_original_name"], "Google Sheet - test-sheet.csv")
        self.assertEqual(session["import_csv_headers"], ["Company Name", "Email"])

    @patch("crm.services.google_sheets.fetch_google_sheet_rows")
    def test_google_sheets_preview_shows_user_friendly_fetch_error(self, mock_fetch):
        mock_fetch.side_effect = RuntimeError("Failed to fetch CSV data from Google Sheets: boom")

        response = self.client.post(
            reverse("import_google_sheets"),
            {
                "sheet_url": "https://docs.google.com/spreadsheets/d/test-sheet/edit?gid=0#gid=0",
                "action": "preview",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Failed to fetch CSV data from Google Sheets: boom")


class BrandingMediaTests(CRMRoleTestMixin, TestCase):
    def setUp(self):
        self.client = Client()
        self.owner_user = self.create_user("owner", role=ROLE_OWNER)
        self.temp_media_root = tempfile.mkdtemp()
        self.settings_override = override_settings(
            DEBUG=False,
            ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"],
            MEDIA_ROOT=self.temp_media_root,
            MEDIA_URL="/media/",
        )
        self.settings_override.enable()
        self.addCleanup(self.settings_override.disable)
        self.addCleanup(shutil.rmtree, self.temp_media_root, ignore_errors=True)

    def test_branding_falls_back_to_site_name_when_no_logo_exists(self):
        SiteBranding.objects.create(site_name="The Zulfis CRM")

        response = self.client.get(reverse("login"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "The Zulfis CRM")
        self.assertNotContains(response, 'class="brand-logo"')

    def test_branding_image_renders_on_login_page(self):
        SiteBranding.objects.create(
            site_name="The Zulfis CRM",
            logo_image=make_logo_file(),
            logo_alt_text="The Zulfis logo",
        )

        response = self.client.get(reverse("login"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="brand-logo"')
        self.assertContains(response, "/media/branding/logo")
        self.assertContains(response, 'alt="The Zulfis logo"')

    def test_site_branding_admin_accepts_uploaded_logo_and_shows_preview(self):
        branding = SiteBranding.objects.create(site_name="The Zulfis CRM")
        self.client.force_login(self.owner_user)

        response = self.client.post(
            reverse("admin:crm_sitebranding_change", args=[branding.pk]),
            {
                "site_name": "The Zulfis CRM",
                "logo_alt_text": "The Zulfis logo",
                "logo_image": make_logo_file("uploaded-logo.png"),
                "_save": "Save",
            },
            follow=True,
        )
        branding.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(branding.logo_image.name.startswith("branding/"))
        self.assertContains(response, "/media/branding/")
        self.assertContains(response, "Logo Preview")

    def test_uploaded_logo_media_url_is_served_with_debug_false(self):
        branding = SiteBranding.objects.create(
            site_name="The Zulfis CRM",
            logo_image=make_logo_file(),
        )

        response = self.client.get(branding.logo_image.url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["Content-Type"], "image/png")

    def test_missing_branding_media_returns_not_found(self):
        response = self.client.get("/media/branding/missing-logo.png")

        self.assertEqual(response.status_code, 404)


class ImportUploadStorageTests(CRMRoleTestMixin, TestCase):
    def setUp(self):
        self.client = Client()
        self.team_lead_user = self.create_user("teamlead", role=ROLE_TEAM_LEAD)
        self.owner_user = self.create_user("owner", role=ROLE_OWNER)
        self.temp_root = Path(tempfile.mkdtemp())
        self.temp_base_dir = self.temp_root / "base"
        self.temp_media_root = self.temp_root / "media"
        self.temp_base_dir.mkdir(parents=True, exist_ok=True)
        self.temp_media_root.mkdir(parents=True, exist_ok=True)
        self.settings_override = override_settings(
            BASE_DIR=self.temp_base_dir,
            MEDIA_ROOT=self.temp_media_root,
            MEDIA_URL="/media/",
        )
        self.settings_override.enable()
        self.addCleanup(self.settings_override.disable)
        self.addCleanup(shutil.rmtree, self.temp_root, ignore_errors=True)

    def test_frontend_import_upload_and_mapping_store_files_under_media_imports(self):
        self.client.force_login(self.team_lead_user)

        upload_response = self.client.post(
            reverse("import_upload"),
            {"csv_file": make_csv_file("frontend-import.csv")},
        )

        self.assertRedirects(upload_response, reverse("import_map_headers"))
        session = self.client.session
        temp_path = Path(session["import_csv_temp_path"])
        self.assertEqual(temp_path.parent, self.temp_media_root / "imports")
        self.assertTrue(temp_path.exists())

        map_response = self.client.post(
            reverse("import_map_headers"),
            {
                "file_name": "frontend-import.csv",
                "map_company_name": "Company Name",
                "map_email": "Email",
            },
            follow=True,
        )
        import_file = ImportFile.objects.get(file_name="frontend-import.csv")

        self.assertEqual(map_response.status_code, 200)
        self.assertEqual(import_file.source_path, str(temp_path))
        self.assertTrue(Path(import_file.source_path).exists())

    def test_admin_import_upload_stores_file_under_media_imports(self):
        self.client.force_login(self.owner_user)

        response = self.client.post(
            reverse("admin:crm_importfile_add"),
            {
                "file_name": "admin-import.csv",
                "csv_file": make_csv_file("admin-import.csv"),
                "rows-TOTAL_FORMS": "0",
                "rows-INITIAL_FORMS": "0",
                "rows-MIN_NUM_FORMS": "0",
                "rows-MAX_NUM_FORMS": "1000",
                "_save": "Save",
            },
            follow=True,
        )
        import_file = ImportFile.objects.get(file_name="admin-import.csv")
        source_path = Path(import_file.source_path)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(source_path.parent, self.temp_media_root / "imports")
        self.assertTrue(source_path.exists())

    def test_legacy_import_uploads_are_copied_to_media_and_paths_are_updated(self):
        legacy_uploads_dir = self.temp_base_dir / "data" / "uploads"
        legacy_uploads_dir.mkdir(parents=True, exist_ok=True)
        legacy_file = legacy_uploads_dir / "legacy-import.csv"
        legacy_file.write_text("Company Name\nAcme Labs\n", encoding="utf-8")
        orphan_file = legacy_uploads_dir / "orphan-import.csv"
        orphan_file.write_text("Company Name\nOrphan Co\n", encoding="utf-8")
        import_file = ImportFile.objects.create(
            file_name="legacy-import.csv",
            source_path=str(legacy_file),
        )

        migration_module = importlib.import_module("crm.migrations.0011_move_import_uploads_to_media")
        migration_module.move_legacy_import_uploads(django_apps, None)
        import_file.refresh_from_db()

        migrated_file = self.temp_media_root / "imports" / legacy_file.name
        migrated_orphan = self.temp_media_root / "imports" / orphan_file.name

        self.assertEqual(import_file.source_path, str(migrated_file))
        self.assertTrue(migrated_file.exists())
        self.assertEqual(
            migrated_file.read_text(encoding="utf-8"),
            legacy_file.read_text(encoding="utf-8"),
        )
        self.assertTrue(migrated_orphan.exists())


@override_settings(
    DEBUG=False,
    ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"],
    SITE_BRAND="The Zulfis CRM",
)
class NotFoundPageTests(CRMRoleTestMixin, TestCase):
    def setUp(self):
        self.client = Client()
        self.staff_user = self.create_user("staffer", role=ROLE_STAFF)

    def test_anonymous_missing_route_uses_branded_404_page(self):
        response = self.client.get("/missing-route/")

        self.assertEqual(response.status_code, 404)
        self.assertTemplateUsed(response, "404.html")
        self.assertContains(response, "Page Not Found | The Zulfis CRM", status_code=404, html=False)
        self.assertContains(response, "The route", status_code=404)
        self.assertContains(response, "/missing-route/", status_code=404)
        self.assertContains(response, 'href="/login/"', status_code=404)

    def test_authenticated_missing_route_shows_companies_recovery_link(self):
        self.client.force_login(self.staff_user)

        response = self.client.get("/still-missing/")

        self.assertEqual(response.status_code, 404)
        self.assertTemplateUsed(response, "404.html")
        self.assertContains(response, "The Zulfis CRM", status_code=404)
        self.assertContains(response, "/still-missing/", status_code=404)
        self.assertContains(response, 'href="/companies/"', status_code=404)
