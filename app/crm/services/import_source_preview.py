"""Preview helpers for import source files."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from crm.models import ImportFile
from crm.services.import_parsers import parse_csv_file, parse_xlsx_file
from crm.services.import_service import get_row_headers


def _normalize_source_type(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        return "xlsx"
    if suffix == ".json":
        return "json"
    return "csv"


def resolve_preview_source(import_file: ImportFile) -> dict[str, object]:
    """Choose the best available source file for preview."""
    original_source_path = getattr(import_file, "original_source_path", "") or ""
    original_source_name = getattr(import_file, "original_source_name", "") or ""
    source_path = getattr(import_file, "source_path", "") or ""
    file_name = getattr(import_file, "file_name", "") or ""

    original_path = Path(original_source_path) if original_source_path else None
    fallback_path = Path(source_path) if source_path else None

    if original_path and original_path.exists():
        filename = original_source_name or original_path.name
        return {
            "available": True,
            "path": original_path,
            "file_name": filename,
            "source_type": _normalize_source_type(filename),
            "is_fallback": False,
            "source_label": "Original upload",
        }

    if fallback_path and fallback_path.exists():
        filename = Path(file_name or fallback_path.name).with_suffix(".csv").name
        return {
            "available": True,
            "path": fallback_path,
            "file_name": filename,
            "source_type": "csv",
            "is_fallback": True,
            "source_label": "Normalized CSV snapshot",
        }

    return {
        "available": False,
        "path": None,
        "file_name": original_source_name or file_name,
        "source_type": "csv",
        "is_fallback": True,
        "source_label": "Unavailable source",
    }


def build_csv_preview(path: str | Path) -> dict[str, object]:
    rows = parse_csv_file(path)
    headers = get_row_headers(rows)
    return {
        "headers": headers,
        "rows": rows,
    }


def build_tabular_preview(
    path: str | Path,
    *,
    source_type: str,
    sheet_name: str | None = None,
) -> dict[str, object]:
    if source_type == "xlsx":
        preview = build_xlsx_preview(path, sheet_name=sheet_name)
    else:
        preview = build_csv_preview(path)

    preview.setdefault("sheet_names", [])
    preview.setdefault("selected_sheet", None)
    preview["row_count"] = len(preview["rows"])
    return preview


def build_json_preview(path: str | Path) -> dict[str, object]:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    return {
        "formatted_json": json.dumps(payload, indent=2, ensure_ascii=False),
    }


def list_xlsx_sheets(path: str | Path) -> list[str]:
    workbook = load_workbook(filename=Path(path), read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def build_xlsx_preview(path: str | Path, *, sheet_name: str | None = None) -> dict[str, object]:
    sheets = list_xlsx_sheets(path)
    selected_sheet = sheet_name if sheet_name in sheets else (sheets[0] if sheets else None)
    rows = parse_xlsx_file(path, sheet_name=selected_sheet) if selected_sheet else []
    headers = get_row_headers(rows)
    return {
        "sheet_names": sheets,
        "selected_sheet": selected_sheet,
        "headers": headers,
        "rows": rows,
    }


def filter_tabular_preview_rows(
    rows: list[dict[str, object]],
    headers: list[str],
    query: str,
) -> list[dict[str, object]]:
    needle = (query or "").strip().casefold()
    if not needle:
        return rows

    filtered_rows: list[dict[str, object]] = []
    keys = headers or (list(rows[0].keys()) if rows else [])
    for row in rows:
        haystack = " ".join(str(row.get(key, "") or "") for key in keys).casefold()
        if needle in haystack:
            filtered_rows.append(row)
    return filtered_rows
