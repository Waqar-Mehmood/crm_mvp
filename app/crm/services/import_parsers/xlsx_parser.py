"""XLSX parser helpers for import sources."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def _normalize_cell(value: object) -> str:
    """Normalize workbook values to strings."""
    return "" if value is None else str(value)


def parse_xlsx_file(xlsx_path: str | Path, sheet_name: str | None = None) -> list[dict[str, str]]:
    """Parse an XLSX file into row dictionaries.

    The active sheet is used by default. This stays intentionally minimal and
    only supports a header row followed by data rows.
    """
    path = Path(xlsx_path)
    workbook = load_workbook(filename=path, read_only=True, data_only=True)

    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []

    headers = [_normalize_cell(cell).strip() for cell in rows[0]]
    if not any(headers):
        return []

    parsed_rows: list[dict[str, str]] = []
    for values in rows[1:]:
        row = {
            header: _normalize_cell(value)
            for header, value in zip(headers, values)
            if header
        }
        if row:
            parsed_rows.append(row)

    return parsed_rows
